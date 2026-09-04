from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, or_
from typing import Optional, List
import pandas as pd
import io
from datetime import datetime, date
import math
import numpy as np
from dateutil import parser

from app.db.session import get_db
from app.schemas.main_db import (
    MainDBCreate,
    MainDBUpdate,
    MainDBResponse,
    MainDBListResponse,
    MainDBSummary,
    ApplicationLogResponse,
)
from app.crud import main_db as crud
from app.crud.main_db import (
    get_main_db_records,
    get_application_logs,
    get_upload_history_paginated,
)
from app.models.main_db import MainDB

from app.models.application_logs import ApplicationLogs
from app.core.deps import get_current_active_user
from app.models.user import User

router = APIRouter(
    prefix="/api/main-db",
    tags=["Main Database"],
    dependencies=[Depends(get_current_active_user)],
)

# ---------------------
# Constants
# ---------------------
COLUMN_MAPPING = {
    "DTN": "DB_DTN",
    "Est. Category": "DB_EST_CAT",
    "LTO Company": "DB_EST_LTO_COMP",
    "LTO Address": "DB_EST_LTO_ADD",
    "Email": "DB_EST_EADD",
    "TIN": "DB_EST_TIN",
    "Contact No.": "DB_EST_CONTACT_NO",
    "LTO No.": "DB_EST_LTO_NO",
    "Validity": "DB_EST_VALIDITY",
    "Brand Name": "DB_PROD_BR_NAME",
    "Generic Name": "DB_PROD_GEN_NAME",
    "Dosage Strength": "DB_PROD_DOS_STR",
    "Dosage Form": "DB_PROD_DOS_FORM",
    "Prescription": "DB_PROD_CLASS_PRESCRIP",
    "Essential Drug": "DB_PROD_ESS_DRUG_LIST",
    "Pharma Category": "DB_PROD_PHARMA_CAT",
    "Manufacturer": "DB_PROD_MANU",
    "Manufacturer Address": "DB_PROD_MANU_ADD",
    "Manufacturer TIN": "DB_PROD_MANU_TIN",
    "Manufacturer LTO No.": "DB_PROD_MANU_LTO_NO",
    "Manufacturer Country": "DB_PROD_MANU_COUNTRY",
    "Trader": "DB_PROD_TRADER",
    "Trader Address": "DB_PROD_TRADER_ADD",
    "Trader TIN": "DB_PROD_TRADER_TIN",
    "Trader LTO No.": "DB_PROD_TRADER_LTO_NO",
    "Trader Country": "DB_PROD_TRADER_COUNTRY",
    "Repacker": "DB_PROD_REPACKER",
    "Repacker Address": "DB_PROD_REPACKER_ADD",
    "Repacker TIN": "DB_PROD_REPACKER_TIN",
    "Repacker LTO No.": "DB_PROD_REPACKER_LTO_NO",
    "Repacker Country": "DB_PROD_REPACKER_COUNTRY",
    "Importer": "DB_PROD_IMPORTER",
    "Importer Address": "DB_PROD_IMPORTER_ADD",
    "Importer TIN": "DB_PROD_IMPORTER_TIN",
    "Importer LTO No.": "DB_PROD_IMPORTER_LTO_NO",
    "Importer Country": "DB_PROD_IMPORTER_COUNTRY",
    "Distributor": "DB_PROD_DISTRI",
    "Distributor Address": "DB_PROD_DISTRI_ADD",
    "Distributor TIN": "DB_PROD_DISTRI_TIN",
    "Distributor LTO No.": "DB_PROD_DISTRI_LTO_NO",
    "Distributor Country": "DB_PROD_DISTRI_COUNTRY",
    "Shelf Life": "DB_PROD_DISTRI_SHELF_LIFE",
    "Storage Condition": "DB_STORAGE_COND",
    "Packaging": "DB_PACKAGING",
    "Suggested RP": "DB_SUGG_RP",
    "No. Sample": "DB_NO_SAMPLE",
    "Expiry Date": "DB_EXPIRY_DATE",
    "CPR Validity": "DB_CPR_VALIDITY",
    "Registration No.": "DB_REG_NO",
    "App Type": "DB_APP_TYPE",
    "Mother App Type": "DB_MOTHER_APP_TYPE",
    "Old RSN": "DB_OLD_RSN",
    "Amendment 1": "DB_AMMEND1",
    "Amendment 2": "DB_AMMEND2",
    "Amendment 3": "DB_AMMEND3",
    "Product Category": "DB_PROD_CAT",
    "Certification": "DB_CERTIFICATION",
    "Fee": "DB_FEE",
    "LRF": "DB_LRF",
    "SURC": "DB_SURC",
    "Total": "DB_TOTAL",
    "OR No.": "DB_OR_NO",
    "Date Issued": "DB_DATE_ISSUED",
    "Date Received FDAC": "DB_DATE_RECEIVED_FDAC",
    "Date Received Central": "DB_DATE_RECEIVED_CENT",
    "MO": "DB_MO",
    "File": "DB_FILE",
    "SECPA": "DB_SECPA",
    "SECPA Exp Date": "DB_SECPA_EXP_DATE",
    "SECPA Issued On": "DB_SECPA_ISSUED_ON",
    "Decking Schedule": "DB_DECKING_SCHED",
    "Evaluation": "DB_EVAL",
    "Date Deck": "DB_DATE_DECK",
    "Remarks 1": "DB_REMARKS_1",
    "Date Remarks": "DB_DATE_REMARKS",
    "Class": "DB_CLASS",
    "Date Released": "DB_DATE_RELEASED",
    "Type Doc Released": "DB_TYPE_DOC_RELEASED",
    "Atta Released": "DB_ATTA_RELEASED",
    "CPR Condition": "DB_CPR_COND",
    "CPR Cond Remarks": "DB_CPR_COND_REMARKS",
    "CPR Cond Add Remarks": "DB_CPR_COND_ADD_REMARKS",
    "App Status": "DB_APP_STATUS",
    "Trash": "DB_TRASH",
    "Pharma Prod Cat": "DB_PHARMA_PROD_CAT",
    "Pharma Prod Cat Label": "DB_PHARMA_PROD_CAT_LABEL",
    "Is in PM": "DB_IS_IN_PM",
    "Timeline Citizen Charter": "DB_TIMELINE_CITIZEN_CHARTER",
    "Processing Type": "DB_PROCESSING_TYPE",
}

# ── Column registry (id -> (header_label, getter, width)) ──
EXPORT_COLUMNS = [
    ("processing_type", "Processing Type", lambda r: r.DB_PROCESSING_TYPE, 18),
    ("dtn", "DTN", lambda r: str(r.DB_DTN) if r.DB_DTN else None, 22),
    ("category", "Category", lambda r: r.DB_EST_CAT, 15),
    ("applicant_company", "Applicant Company", lambda r: r.DB_EST_LTO_COMP, 35),
    ("address", "Address", lambda r: r.DB_EST_LTO_ADD, 40),
    ("email", "Email Address", lambda r: r.DB_EST_EADD, 30),
    ("tin", "TIN", lambda r: r.DB_EST_TIN, 15),
    ("contact_no", "Contact No.", lambda r: r.DB_EST_CONTACT_NO, 18),
    ("lto_no", "LTO No.", lambda r: r.DB_EST_LTO_NO, 18),
    ("validity", "Validity", lambda r: r.DB_EST_VALIDITY, 15),
    ("brand_name", "Brand Name", lambda r: r.DB_PROD_BR_NAME, 30),
    ("generic_name", "Generic Name", lambda r: r.DB_PROD_GEN_NAME, 30),
    ("dosage_strength", "Dosage Strength", lambda r: r.DB_PROD_DOS_STR, 20),
    (
        "dosage_form",
        "Dosage Form and Route of Administration",
        lambda r: r.DB_PROD_DOS_FORM,
        30,
    ),
    ("classification", "Classification", lambda r: r.DB_PROD_CLASS_PRESCRIP, 22),
    ("essential_drug", "Essential Drug List", lambda r: r.DB_PROD_ESS_DRUG_LIST, 20),
    ("pharma_cat", "Pharmacologic Category", lambda r: r.DB_PROD_PHARMA_CAT, 25),
    ("manufacturer", "Manufacturer", lambda r: r.DB_PROD_MANU, 30),
    ("manufacturer_add", "Manufacturer Address", lambda r: r.DB_PROD_MANU_ADD, 40),
    ("manufacturer_tin", "Manufacturer TIN", lambda r: r.DB_PROD_MANU_TIN, 15),
    ("manufacturer_lto", "Manufacturer LTO No.", lambda r: r.DB_PROD_MANU_LTO_NO, 20),
    (
        "manufacturer_country",
        "Manufacturer Country",
        lambda r: r.DB_PROD_MANU_COUNTRY,
        20,
    ),
    ("trader", "Trader", lambda r: r.DB_PROD_TRADER, 30),
    ("trader_add", "Trader Address", lambda r: r.DB_PROD_TRADER_ADD, 40),
    ("trader_tin", "Trader TIN", lambda r: r.DB_PROD_TRADER_TIN, 15),
    ("trader_lto", "Trader LTO No.", lambda r: r.DB_PROD_TRADER_LTO_NO, 20),
    ("trader_country", "Trader Country", lambda r: r.DB_PROD_TRADER_COUNTRY, 20),
    ("repacker", "Repacker", lambda r: r.DB_PROD_REPACKER, 30),
    ("repacker_add", "Repacker Address", lambda r: r.DB_PROD_REPACKER_ADD, 40),
    ("repacker_tin", "Repacker TIN", lambda r: r.DB_PROD_REPACKER_TIN, 15),
    ("repacker_lto", "Repacker LTO No.", lambda r: r.DB_PROD_REPACKER_LTO_NO, 20),
    ("repacker_country", "Repacker Country", lambda r: r.DB_PROD_REPACKER_COUNTRY, 20),
    ("importer", "Importer", lambda r: r.DB_PROD_IMPORTER, 30),
    ("importer_add", "Importer Address", lambda r: r.DB_PROD_IMPORTER_ADD, 40),
    ("importer_tin", "Importer TIN", lambda r: r.DB_PROD_IMPORTER_TIN, 15),
    ("importer_lto", "Importer LTO No.", lambda r: r.DB_PROD_IMPORTER_LTO_NO, 20),
    ("importer_country", "Importer Country", lambda r: r.DB_PROD_IMPORTER_COUNTRY, 20),
    ("distributor", "Distributor", lambda r: r.DB_PROD_DISTRI, 30),
    ("distributor_add", "Distributor Address", lambda r: r.DB_PROD_DISTRI_ADD, 40),
    ("distributor_tin", "Distributor TIN", lambda r: r.DB_PROD_DISTRI_TIN, 15),
    ("distributor_lto", "Distributor LTO No.", lambda r: r.DB_PROD_DISTRI_LTO_NO, 20),
    (
        "distributor_country",
        "Distributor Country",
        lambda r: r.DB_PROD_DISTRI_COUNTRY,
        20,
    ),
    ("shelf_life", "Shelf Life", lambda r: r.DB_PROD_DISTRI_SHELF_LIFE, 20),
    ("storage_cond", "Storage Condition", lambda r: r.DB_STORAGE_COND, 25),
    ("packaging", "Packaging", lambda r: r.DB_PACKAGING, 25),
    ("suggested_rp", "Suggested Retail Price", lambda r: r.DB_SUGG_RP, 20),
    ("reg_no", "Registration Number", lambda r: r.DB_REG_NO, 22),
    ("app_type", "Application Type", lambda r: r.DB_APP_TYPE, 20),
    ("mother_app_type", "Mother Application Type", lambda r: r.DB_MOTHER_APP_TYPE, 25),
    ("old_rsn", "Old RSN/ Other DTN", lambda r: r.DB_OLD_RSN, 20),
    ("amendment1", "Amendment 1", lambda r: r.DB_AMMEND1, 15),
    ("amendment2", "Amendment 2", lambda r: r.DB_AMMEND2, 15),
    ("amendment3", "Amendment 3", lambda r: r.DB_AMMEND3, 15),
    ("prod_cat", "Product Category", lambda r: r.DB_PROD_CAT, 20),
    ("certification", "Certification", lambda r: r.DB_CERTIFICATION, 20),
    ("fee", "Fee", lambda r: r.DB_FEE, 12),
    ("lrf", "LRF", lambda r: r.DB_LRF, 12),
    ("surc", "SURC", lambda r: r.DB_SURC, 12),
    ("total", "Total", lambda r: r.DB_TOTAL, 12),
    ("or_no", "OR No.", lambda r: r.DB_OR_NO, 18),
    ("date_issued", "Date Issued", lambda r: r.DB_DATE_ISSUED, 18),
    (
        "date_received_fdac",
        "Date when the application was received by FDAC",
        lambda r: r.DB_DATE_RECEIVED_FDAC,
        35,
    ),
    (
        "date_received_cent",
        "Date when the application was received by CDRR",
        lambda r: r.DB_DATE_RECEIVED_CENT,
        35,
    ),
    ("mo", "MO", lambda r: r.DB_MO, 12),
    ("file_copy", "FILE COPY", lambda r: r.DB_FILE, 12),
    ("secpa", "SECPA No.", lambda r: r.DB_SECPA, 18),
    ("secpa_exp", "Expiry Date", lambda r: r.DB_SECPA_EXP_DATE, 18),
    ("secpa_issued", "Issued On", lambda r: r.DB_SECPA_ISSUED_ON, 18),
    (
        "remarks1",
        "Remarks (1)(e.g. reason of application returned)",
        lambda r: r.DB_REMARKS_1,
        45,
    ),
    ("date_remarks", "Date of Remarks (1)", lambda r: r.DB_DATE_REMARKS, 20),
    ("class", "Class", lambda r: r.DB_CLASS, 15),
    ("date_released", "Date Released by the CDRR", lambda r: r.DB_DATE_RELEASED, 25),
    (
        "type_doc_released",
        "Type of Document Released",
        lambda r: r.DB_TYPE_DOC_RELEASED,
        25,
    ),
    (
        "atta_released",
        "Attachment/s released with authorization",
        lambda r: r.DB_ATTA_RELEASED,
        35,
    ),
    (
        "cpr_cond",
        "CPR Condition/s Ticked at the back of CPR",
        lambda r: r.DB_CPR_COND,
        40,
    ),
    ("cpr_cond_remarks", "CPR Cond Remarks", lambda r: r.DB_CPR_COND_REMARKS, 30),
    (
        "cpr_cond_add_remarks",
        "CPR Cond Additional Remarks",
        lambda r: r.DB_CPR_COND_ADD_REMARKS,
        35,
    ),
    ("app_status", "App Status", lambda r: r.DB_APP_STATUS, 20),
    ("app_remarks", "App Remarks", lambda r: r.DB_APP_REMARKS, 30),
    ("timeline", "Timeline (Days)", lambda r: r.DB_TIMELINE_CITIZEN_CHARTER, 18),
]


# ── LOG_STEPS as "column groups" — each step = 1 selectable id, expands to 4 real columns ──
def _log_step_export_id(step_label):
    return f"log__{step_label.lower().replace(' ', '_').replace('&', 'and')}"


# ============================================================
# UPDATED LOG_STEPS CONFIGURATION
# ============================================================
# Format: (step_label, user_col, decision_col, remarks_col, date_col, thread_col, del_idx)
# Format: (step_label, user_col, id_col, decision_col, remarks_col, date_col, thread_col, del_idx)

LOG_STEPS = [
    (
        "Decking",
        "Decker",
        "Decker ID",
        "Decker Decision",
        "Decker Remarks",
        "Date Decked End",
        "Decker Del Thread",
        1,
    ),
    (
        "Quality Evaluation",
        "Evaluator",
        "Evaluator ID",
        "Evaluator Decision",
        "Evaluator Remarks",
        "Date Eval End",
        "Evaluator Del Thread",
        2,
    ),
    (
        "S&E",
        "S&E",
        "S&E ID",
        "S&E Decision",
        "S&E Remarks",
        "Date S&E End",
        "S&E Del Thread",
        2,
    ),
    (
        "Checking",
        "Checker",
        "Checker ID",
        "Checker Decision",
        "Checker Remarks",
        "Date Checker End",
        "Checker Del Thread",
        3,
    ),
    (
        "Supervisor",
        "Supervisor",
        "Supervisor ID",
        "Supervisor Decision",
        "Supervisor Remarks",
        "Date Supervisor End",
        "Supervisor Del Thread",
        4,
    ),
    (
        "QA Admin",
        "QA Admin",
        "QA Admin ID",
        "QA Admin Decision",
        "QA Admin Remarks",
        "Date QA Admin End",
        "QA Admin Del Thread",
        5,
    ),
    (
        "LRD Chief Admin",
        "LRD Chief Admin",
        "LRD Chief Admin ID",
        "LRD Chief Admin Decision",
        "LRD Chief Admin Remarks",
        "Date LRD Chief Admin End",
        "LRD Chief Admin Del Thread",
        6,
    ),
    (
        "OD-Receiving",
        "OD-Receiving",
        "OD-Receiving ID",
        "OD-Receiving Decision",
        "OD-Receiving Remarks",
        "Date OD-Receiving End",
        "OD-Receiving Del Thread",
        7,
    ),
    (
        "OD-Releasing",
        "OD-Releasing",
        "OD-Releasing ID",
        "OD-Releasing Decision",
        "OD-Releasing Remarks",
        "Date OD-Releasing End",
        "OD-Releasing Del Thread",
        8,
    ),
    (
        "Releasing Officer",
        "Releasing Officer",
        "Releasing Officer ID",
        "Releasing Officer Decision",
        "Releasing Officer Remarks",
        "Date Releasing Officer End",
        "Releasing Officer Del Thread",
        9,
    ),
]

DATE_FIELDS = {
    "DB_EST_VALIDITY",
    "DB_EXPIRY_DATE",
    "DB_CPR_VALIDITY",
    "DB_DATE_ISSUED",
    "DB_DATE_DECK",
    "DB_DATE_RECEIVED_FDAC",
    "DB_DATE_RECEIVED_CENT",
    "DB_SECPA_EXP_DATE",
    "DB_SECPA_ISSUED_ON",
    "DB_DATE_REMARKS",
    "DB_DATE_RELEASED",
}

NUMERIC_STRING_FIELDS = {"DB_FEE", "DB_LRF", "DB_SURC", "DB_TOTAL"}

INTEGER_FIELDS = {"DB_DTN", "DB_IS_IN_PM", "DB_TIMELINE_CITIZEN_CHARTER"}


# ---------------------
# Helper Functions
# ---------------------
def parse_date_value(value):
    """Parse various date formats and return datetime object or None"""
    if pd.isna(value) or value is None or value == "":
        return None
    if isinstance(value, (datetime, pd.Timestamp)):
        return value
    if isinstance(value, (int, float, np.integer, np.floating)):
        return None
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None
        try:
            parsed_date = parser.parse(value, fuzzy=True)
            return parsed_date
        except:
            return None
    return None


# ---------------------
# Helper: apply processing_type filter to a query
# ---------------------
def _apply_processing_type_filter(query, processing_type):
    """Reusable helper — applies processing_type filter to any SQLAlchemy query."""
    if processing_type is not None:
        if processing_type == "__EMPTY__":
            query = query.filter(
                or_(
                    MainDB.DB_PROCESSING_TYPE.is_(None), MainDB.DB_PROCESSING_TYPE == ""
                )
            )
        else:
            query = query.filter(MainDB.DB_PROCESSING_TYPE == processing_type)
    return query


# ---------------------
# Routes
# ---------------------
@router.get("/", response_model=MainDBListResponse)
def get_main_db(
    page: int = Query(1, ge=1),
    page_size: int = Query(10, ge=1, le=1000),
    search: Optional[str] = Query(None),
    dtns: Optional[str] = Query(
        None,
        description="Comma-separated DTN numbers, e.g. 20260319173438,20260422154843",
    ),
    status: Optional[str] = Query(None),
    category: Optional[str] = Query(None),
    prescription: Optional[str] = Query(None),
    prescription_not: Optional[str] = Query(None),
    dtn: Optional[int] = Query(None),
    manufacturer: Optional[str] = Query(None),
    lto_company: Optional[str] = Query(None),
    brand_name: Optional[str] = Query(None),
    generic_name: Optional[str] = Query(None),
    app_status: Optional[str] = Query(None),
    app_type: Optional[str] = Query(None),
    processing_type: Optional[str] = Query(None),
    entry_type: Optional[str] = Query(None),
    dosage_form: Optional[str] = Query(None),
    manufacturer_country: Optional[str] = Query(None),
    trader: Optional[str] = Query(None),
    trader_country: Optional[str] = Query(None),
    importer: Optional[str] = Query(None),
    importer_country: Optional[str] = Query(None),
    distributor: Optional[str] = Query(None),
    distributor_country: Optional[str] = Query(None),
    repacker: Optional[str] = Query(None),
    repacker_country: Optional[str] = Query(None),
    type_doc_released: Optional[str] = Query(None),
    date_released_from: Optional[str] = Query(None),
    date_released_to: Optional[str] = Query(None),
    date_received_cent_from: Optional[str] = Query(None),
    date_received_cent_to: Optional[str] = Query(None),
    user_uploader: Optional[str] = Query(None),
    date_excel_upload_from: Optional[str] = Query(None),
    date_excel_upload_to: Optional[str] = Query(None),
    null_date_released: Optional[str] = Query(None),
    null_date_received_cent: Optional[str] = Query(None),
    sort_by: str = Query("DB_DATE_EXCEL_UPLOAD"),
    sort_order: str = Query("desc", pattern="^(asc|desc)$"),
    db: Session = Depends(get_db),
):
    """Get paginated list of main database records with flexible filtering"""
    skip = (page - 1) * page_size

    # ✅ BAGO — parse comma-separated DTNs into list of ints
    parsed_dtns = []
    if dtns:
        for raw in dtns.split(","):
            raw = raw.strip()
            if raw.isdigit():
                parsed_dtns.append(int(raw))

    filters = {
        "status": status,
        "category": category,
        "prescription": prescription,
        "prescription_not": prescription_not,
        "dtn": dtn,
        "dtns": parsed_dtns if parsed_dtns else None,  # ✅ BAGO
        "manufacturer": manufacturer,
        "lto_company": lto_company,
        "brand_name": brand_name,
        "generic_name": generic_name,
        "app_status": app_status,
        "app_type": app_type,
        "processing_type": processing_type,
        "entry_type": entry_type,
        "dosage_form": dosage_form,
        "manufacturer_country": manufacturer_country,
        "trader": trader,
        "trader_country": trader_country,
        "importer": importer,
        "importer_country": importer_country,
        "distributor": distributor,
        "distributor_country": distributor_country,
        "repacker": repacker,
        "repacker_country": repacker_country,
        "type_doc_released": type_doc_released,
        "date_released_from": date_released_from,
        "date_released_to": date_released_to,
        "date_received_cent_from": date_received_cent_from,
        "date_received_cent_to": date_received_cent_to,
        "user_uploader": user_uploader,
        "null_date_released": null_date_released,
        "null_date_received_cent": null_date_received_cent,
        "date_excel_upload_from": date_excel_upload_from,
        "date_excel_upload_to": date_excel_upload_to,
    }

    records, total = get_main_db_records(
        db=db,
        skip=skip,
        limit=page_size,
        search=search,
        filters=filters,
        sort_by=sort_by,
        sort_order=sort_order,
    )
    total_pages = math.ceil(total / page_size) if total > 0 else 0
    return {
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": total_pages,
        "data": records,
    }


# ✅ Get unique processing types with counts
@router.get("/processing-types")
def get_processing_types(
    status: Optional[str] = Query(
        None, description="Filter by decking status: 'not_decked' or 'decked'"
    ),
    app_type: Optional[str] = Query(None, description="Filter by application type"),
    prescription: Optional[str] = Query(
        None, description="Filter by prescription type"
    ),
    app_status: Optional[str] = Query(None, description="Filter by app status"),
    db: Session = Depends(get_db),
):
    """Get unique DB_PROCESSING_TYPE values with counts, filtered by active sidebar selections"""

    def _build_base_query(base_query):
        if status == "decked":
            _decked_ids = (
                db.query(ApplicationLogs.main_db_id)
                .filter(ApplicationLogs.application_step == "Decking")
                .subquery()
            )
            base_query = base_query.filter(
                or_(MainDB.DB_ID.in_(_decked_ids), MainDB.DB_APP_STATUS == "Completed")
            )
        elif status == "not_decked":
            _decked_ids = (
                db.query(ApplicationLogs.main_db_id)
                .filter(ApplicationLogs.application_step == "Decking")
                .subquery()
            )
            base_query = base_query.filter(
                MainDB.DB_ID.notin_(_decked_ids),
                or_(
                    MainDB.DB_APP_STATUS.is_(None),
                    MainDB.DB_APP_STATUS == "",
                    MainDB.DB_APP_STATUS != "Completed",
                ),
            )

        if app_type is not None:
            if app_type == "__EMPTY__":
                base_query = base_query.filter(
                    or_(MainDB.DB_APP_TYPE.is_(None), MainDB.DB_APP_TYPE == "")
                )
            else:
                base_query = base_query.filter(MainDB.DB_APP_TYPE == app_type)

        if prescription is not None:
            if prescription == "__EMPTY__":
                base_query = base_query.filter(
                    or_(
                        MainDB.DB_PROD_CLASS_PRESCRIP.is_(None),
                        MainDB.DB_PROD_CLASS_PRESCRIP == "",
                    )
                )
            else:
                base_query = base_query.filter(
                    MainDB.DB_PROD_CLASS_PRESCRIP == prescription
                )

        if app_status is not None:
            if app_status == "__EMPTY__":
                base_query = base_query.filter(
                    or_(MainDB.DB_APP_STATUS.is_(None), MainDB.DB_APP_STATUS == "")
                )
            else:
                base_query = base_query.filter(MainDB.DB_APP_STATUS == app_status)

        return base_query

    query = _build_base_query(
        db.query(MainDB.DB_PROCESSING_TYPE, func.count(MainDB.DB_ID).label("count"))
    )
    results = (
        query.filter(
            MainDB.DB_PROCESSING_TYPE.isnot(None), MainDB.DB_PROCESSING_TYPE != ""
        )
        .group_by(MainDB.DB_PROCESSING_TYPE)
        .order_by(MainDB.DB_PROCESSING_TYPE)
        .all()
    )

    query_no_type = _build_base_query(db.query(func.count(MainDB.DB_ID)))
    no_type_count = query_no_type.filter(
        or_(MainDB.DB_PROCESSING_TYPE.is_(None), MainDB.DB_PROCESSING_TYPE == "")
    ).scalar()

    processing_types = [{"value": pt, "count": count} for pt, count in results]

    if no_type_count and no_type_count > 0:
        processing_types.insert(0, {"value": None, "count": no_type_count})

    return {"processing_types": processing_types}


@router.get("/app-types")
def get_app_types(
    status: Optional[str] = Query(
        None, description="Filter by decking status: 'not_decked' or 'decked'"
    ),
    processing_type: Optional[str] = Query(
        None, description="Filter by processing type"
    ),
    prescription: Optional[str] = Query(
        None, description="Filter by prescription/classification"
    ),
    app_status: Optional[str] = Query(None, description="Filter by app status"),
    db: Session = Depends(get_db),
):
    """Get unique DB_APP_TYPE values with counts"""
    query = db.query(MainDB.DB_APP_TYPE, func.count(MainDB.DB_ID).label("count"))

    if status == "decked":
        _decked_ids = (
            db.query(ApplicationLogs.main_db_id)
            .filter(ApplicationLogs.application_step == "Decking")
            .subquery()
        )
        query = query.filter(
            or_(MainDB.DB_ID.in_(_decked_ids), MainDB.DB_APP_STATUS == "Completed")
        )
    elif status == "not_decked":
        _decked_ids = (
            db.query(ApplicationLogs.main_db_id)
            .filter(ApplicationLogs.application_step == "Decking")
            .subquery()
        )
        query = query.filter(
            MainDB.DB_ID.notin_(_decked_ids),
            or_(
                MainDB.DB_APP_STATUS.is_(None),
                MainDB.DB_APP_STATUS == "",
                MainDB.DB_APP_STATUS != "Completed",
            ),
        )

    query = _apply_processing_type_filter(query, processing_type)

    if prescription is not None:
        if prescription == "__EMPTY__" or prescription == "":
            query = query.filter(
                or_(
                    MainDB.DB_PROD_CLASS_PRESCRIP.is_(None),
                    MainDB.DB_PROD_CLASS_PRESCRIP == "",
                )
            )
        else:
            query = query.filter(MainDB.DB_PROD_CLASS_PRESCRIP == prescription)

    if app_status is not None:
        if app_status == "__EMPTY__" or app_status == "":
            query = query.filter(
                or_(MainDB.DB_APP_STATUS.is_(None), MainDB.DB_APP_STATUS == "")
            )
        else:
            query = query.filter(MainDB.DB_APP_STATUS == app_status)

    results_with_type = (
        query.filter(MainDB.DB_APP_TYPE.isnot(None), MainDB.DB_APP_TYPE != "")
        .group_by(MainDB.DB_APP_TYPE)
        .order_by(MainDB.DB_APP_TYPE)
        .all()
    )

    query_no_type = db.query(func.count(MainDB.DB_ID))

    if status == "decked":
        _decked_ids = (
            db.query(ApplicationLogs.main_db_id)
            .filter(ApplicationLogs.application_step == "Decking")
            .subquery()
        )
        query_no_type = query_no_type.filter(
            or_(MainDB.DB_ID.in_(_decked_ids), MainDB.DB_APP_STATUS == "Completed")
        )
    elif status == "not_decked":
        _decked_ids = (
            db.query(ApplicationLogs.main_db_id)
            .filter(ApplicationLogs.application_step == "Decking")
            .subquery()
        )
        query_no_type = query_no_type.filter(
            MainDB.DB_ID.notin_(_decked_ids),
            or_(
                MainDB.DB_APP_STATUS.is_(None),
                MainDB.DB_APP_STATUS == "",
                MainDB.DB_APP_STATUS != "Completed",
            ),
        )

    query_no_type = _apply_processing_type_filter(query_no_type, processing_type)

    if prescription is not None:
        if prescription == "__EMPTY__" or prescription == "":
            query_no_type = query_no_type.filter(
                or_(
                    MainDB.DB_PROD_CLASS_PRESCRIP.is_(None),
                    MainDB.DB_PROD_CLASS_PRESCRIP == "",
                )
            )
        else:
            query_no_type = query_no_type.filter(
                MainDB.DB_PROD_CLASS_PRESCRIP == prescription
            )

    if app_status is not None:
        if app_status == "__EMPTY__" or app_status == "":
            query_no_type = query_no_type.filter(
                or_(MainDB.DB_APP_STATUS.is_(None), MainDB.DB_APP_STATUS == "")
            )
        else:
            query_no_type = query_no_type.filter(MainDB.DB_APP_STATUS == app_status)

    no_type_count = query_no_type.filter(
        or_(MainDB.DB_APP_TYPE.is_(None), MainDB.DB_APP_TYPE == "")
    ).scalar()

    app_types = [
        {"value": app_type, "count": count} for app_type, count in results_with_type
    ]

    if no_type_count and no_type_count > 0:
        app_types.insert(0, {"value": None, "count": no_type_count})

    return {"app_types": app_types}


@router.get("/prescription-types")
def get_prescription_types(
    status: Optional[str] = Query(
        None, description="Filter by decking status: 'not_decked' or 'decked'"
    ),
    app_type: Optional[str] = Query(None, description="Filter by application type"),
    processing_type: Optional[str] = Query(
        None, description="Filter by processing type"
    ),
    app_status: Optional[str] = Query(None, description="Filter by app status"),
    db: Session = Depends(get_db),
):
    """Get unique DB_PROD_CLASS_PRESCRIP values with counts"""
    query = db.query(
        MainDB.DB_PROD_CLASS_PRESCRIP, func.count(MainDB.DB_ID).label("count")
    )

    if status == "decked":
        _decked_ids = (
            db.query(ApplicationLogs.main_db_id)
            .filter(ApplicationLogs.application_step == "Decking")
            .subquery()
        )
        query = query.filter(
            or_(MainDB.DB_ID.in_(_decked_ids), MainDB.DB_APP_STATUS == "Completed")
        )
    elif status == "not_decked":
        _decked_ids = (
            db.query(ApplicationLogs.main_db_id)
            .filter(ApplicationLogs.application_step == "Decking")
            .subquery()
        )
        query = query.filter(
            MainDB.DB_ID.notin_(_decked_ids),
            or_(
                MainDB.DB_APP_STATUS.is_(None),
                MainDB.DB_APP_STATUS == "",
                MainDB.DB_APP_STATUS != "Completed",
            ),
        )

    if app_type is not None:
        if app_type == "__EMPTY__" or app_type == "":
            query = query.filter(
                or_(MainDB.DB_APP_TYPE.is_(None), MainDB.DB_APP_TYPE == "")
            )
        else:
            query = query.filter(MainDB.DB_APP_TYPE == app_type)

    query = _apply_processing_type_filter(query, processing_type)

    if app_status is not None:
        if app_status == "__EMPTY__" or app_status == "":
            query = query.filter(
                or_(MainDB.DB_APP_STATUS.is_(None), MainDB.DB_APP_STATUS == "")
            )
        else:
            query = query.filter(MainDB.DB_APP_STATUS == app_status)

    results_with_type = (
        query.filter(
            MainDB.DB_PROD_CLASS_PRESCRIP.isnot(None),
            MainDB.DB_PROD_CLASS_PRESCRIP != "",
        )
        .group_by(MainDB.DB_PROD_CLASS_PRESCRIP)
        .order_by(MainDB.DB_PROD_CLASS_PRESCRIP)
        .all()
    )

    query_no_type = db.query(func.count(MainDB.DB_ID))

    if status == "decked":
        _decked_ids = (
            db.query(ApplicationLogs.main_db_id)
            .filter(ApplicationLogs.application_step == "Decking")
            .subquery()
        )
        query_no_type = query_no_type.filter(
            or_(MainDB.DB_ID.in_(_decked_ids), MainDB.DB_APP_STATUS == "Completed")
        )
    elif status == "not_decked":
        _decked_ids = (
            db.query(ApplicationLogs.main_db_id)
            .filter(ApplicationLogs.application_step == "Decking")
            .subquery()
        )
        query_no_type = query_no_type.filter(
            MainDB.DB_ID.notin_(_decked_ids),
            or_(
                MainDB.DB_APP_STATUS.is_(None),
                MainDB.DB_APP_STATUS == "",
                MainDB.DB_APP_STATUS != "Completed",
            ),
        )

    if app_type is not None:
        if app_type == "__EMPTY__" or app_type == "":
            query_no_type = query_no_type.filter(
                or_(MainDB.DB_APP_TYPE.is_(None), MainDB.DB_APP_TYPE == "")
            )
        else:
            query_no_type = query_no_type.filter(MainDB.DB_APP_TYPE == app_type)

    query_no_type = _apply_processing_type_filter(query_no_type, processing_type)

    if app_status is not None:
        if app_status == "__EMPTY__" or app_status == "":
            query_no_type = query_no_type.filter(
                or_(MainDB.DB_APP_STATUS.is_(None), MainDB.DB_APP_STATUS == "")
            )
        else:
            query_no_type = query_no_type.filter(MainDB.DB_APP_STATUS == app_status)

    no_type_count = query_no_type.filter(
        or_(
            MainDB.DB_PROD_CLASS_PRESCRIP.is_(None), MainDB.DB_PROD_CLASS_PRESCRIP == ""
        )
    ).scalar()

    prescription_types = [
        {"value": pres_type, "count": count} for pres_type, count in results_with_type
    ]

    if no_type_count and no_type_count > 0:
        prescription_types.insert(0, {"value": None, "count": no_type_count})

    return {"prescription_types": prescription_types}


@router.get("/app-status-types")
def get_app_status_types(
    status: Optional[str] = Query(
        None, description="Filter by decking status: 'not_decked' or 'decked'"
    ),
    app_type: Optional[str] = Query(None, description="Filter by application type"),
    prescription: Optional[str] = Query(
        None, description="Filter by prescription type"
    ),
    processing_type: Optional[str] = Query(
        None, description="Filter by processing type"
    ),
    db: Session = Depends(get_db),
):
    """Get unique DB_APP_STATUS values with counts"""
    query = db.query(MainDB.DB_APP_STATUS, func.count(MainDB.DB_ID).label("count"))

    if status == "decked":
        _decked_ids = (
            db.query(ApplicationLogs.main_db_id)
            .filter(ApplicationLogs.application_step == "Decking")
            .subquery()
        )
        query = query.filter(
            or_(MainDB.DB_ID.in_(_decked_ids), MainDB.DB_APP_STATUS == "Completed")
        )
    elif status == "not_decked":
        _decked_ids = (
            db.query(ApplicationLogs.main_db_id)
            .filter(ApplicationLogs.application_step == "Decking")
            .subquery()
        )
        query = query.filter(
            MainDB.DB_ID.notin_(_decked_ids),
            or_(
                MainDB.DB_APP_STATUS.is_(None),
                MainDB.DB_APP_STATUS == "",
                MainDB.DB_APP_STATUS != "Completed",
            ),
        )

    if app_type is not None:
        if app_type == "__EMPTY__" or app_type == "":
            query = query.filter(
                or_(MainDB.DB_APP_TYPE.is_(None), MainDB.DB_APP_TYPE == "")
            )
        else:
            query = query.filter(MainDB.DB_APP_TYPE == app_type)

    if prescription is not None:
        if prescription == "__EMPTY__" or prescription == "":
            query = query.filter(
                or_(
                    MainDB.DB_PROD_CLASS_PRESCRIP.is_(None),
                    MainDB.DB_PROD_CLASS_PRESCRIP == "",
                )
            )
        else:
            query = query.filter(MainDB.DB_PROD_CLASS_PRESCRIP == prescription)

    query = _apply_processing_type_filter(query, processing_type)

    results_with_status = (
        query.filter(MainDB.DB_APP_STATUS.isnot(None), MainDB.DB_APP_STATUS != "")
        .group_by(MainDB.DB_APP_STATUS)
        .order_by(MainDB.DB_APP_STATUS)
        .all()
    )

    query_no_status = db.query(func.count(MainDB.DB_ID))

    if status == "decked":
        _decked_ids = (
            db.query(ApplicationLogs.main_db_id)
            .filter(ApplicationLogs.application_step == "Decking")
            .subquery()
        )
        query_no_status = query_no_status.filter(
            or_(MainDB.DB_ID.in_(_decked_ids), MainDB.DB_APP_STATUS == "Completed")
        )
    elif status == "not_decked":
        _decked_ids = (
            db.query(ApplicationLogs.main_db_id)
            .filter(ApplicationLogs.application_step == "Decking")
            .subquery()
        )
        query_no_status = query_no_status.filter(
            MainDB.DB_ID.notin_(_decked_ids),
            or_(
                MainDB.DB_APP_STATUS.is_(None),
                MainDB.DB_APP_STATUS == "",
                MainDB.DB_APP_STATUS != "Completed",
            ),
        )

    if app_type is not None:
        if app_type == "__EMPTY__" or app_type == "":
            query_no_status = query_no_status.filter(
                or_(MainDB.DB_APP_TYPE.is_(None), MainDB.DB_APP_TYPE == "")
            )
        else:
            query_no_status = query_no_status.filter(MainDB.DB_APP_TYPE == app_type)

    if prescription is not None:
        if prescription == "__EMPTY__" or prescription == "":
            query_no_status = query_no_status.filter(
                or_(
                    MainDB.DB_PROD_CLASS_PRESCRIP.is_(None),
                    MainDB.DB_PROD_CLASS_PRESCRIP == "",
                )
            )
        else:
            query_no_status = query_no_status.filter(
                MainDB.DB_PROD_CLASS_PRESCRIP == prescription
            )

    query_no_status = _apply_processing_type_filter(query_no_status, processing_type)

    no_status_count = query_no_status.filter(
        or_(MainDB.DB_APP_STATUS.is_(None), MainDB.DB_APP_STATUS == "")
    ).scalar()

    app_status_types = [
        {"value": app_status, "count": count}
        for app_status, count in results_with_status
    ]

    if no_status_count and no_status_count > 0:
        app_status_types.insert(0, {"value": None, "count": no_status_count})

    return {"app_status_types": app_status_types}


@router.get("/establishment-categories")
def get_establishment_categories(
    status: Optional[str] = Query(
        None,
        description="Filter by decking status: 'not_decked', 'decked', or null for all",
    ),
    db: Session = Depends(get_db),
):
    """Get unique DB_EST_CAT (Establishment Category) values with counts"""
    query = db.query(MainDB.DB_EST_CAT, func.count(MainDB.DB_ID).label("count"))

    if status == "decked":
        _decked_ids = (
            db.query(ApplicationLogs.main_db_id)
            .filter(ApplicationLogs.application_step == "Decking")
            .subquery()
        )
        query = query.filter(
            or_(MainDB.DB_ID.in_(_decked_ids), MainDB.DB_APP_STATUS == "Completed")
        )
    elif status == "not_decked":
        _decked_ids = (
            db.query(ApplicationLogs.main_db_id)
            .filter(ApplicationLogs.application_step == "Decking")
            .subquery()
        )
        query = query.filter(
            MainDB.DB_ID.notin_(_decked_ids),
            or_(
                MainDB.DB_APP_STATUS.is_(None),
                MainDB.DB_APP_STATUS == "",
                MainDB.DB_APP_STATUS != "Completed",
            ),
        )

    results_with_category = (
        query.filter(MainDB.DB_EST_CAT.isnot(None), MainDB.DB_EST_CAT != "")
        .group_by(MainDB.DB_EST_CAT)
        .order_by(MainDB.DB_EST_CAT)
        .all()
    )

    query_no_category = db.query(func.count(MainDB.DB_ID))

    if status == "decked":
        _decked_ids = (
            db.query(ApplicationLogs.main_db_id)
            .filter(ApplicationLogs.application_step == "Decking")
            .subquery()
        )
        query_no_category = query_no_category.filter(
            or_(MainDB.DB_ID.in_(_decked_ids), MainDB.DB_APP_STATUS == "Completed")
        )
    elif status == "not_decked":
        _decked_ids = (
            db.query(ApplicationLogs.main_db_id)
            .filter(ApplicationLogs.application_step == "Decking")
            .subquery()
        )
        query_no_category = query_no_category.filter(
            MainDB.DB_ID.notin_(_decked_ids),
            or_(
                MainDB.DB_APP_STATUS.is_(None),
                MainDB.DB_APP_STATUS == "",
                MainDB.DB_APP_STATUS != "Completed",
            ),
        )

    no_category_count = query_no_category.filter(
        or_(MainDB.DB_EST_CAT.is_(None), MainDB.DB_EST_CAT == "")
    ).scalar()

    categories = [
        {"value": category, "count": count} for category, count in results_with_category
    ]

    if no_category_count and no_category_count > 0:
        categories.insert(0, {"value": None, "count": no_category_count})

    return {"categories": categories}


@router.get("/entry-types")
def get_entry_types(
    status: Optional[str] = Query(
        None, description="Filter by decking status: 'not_decked' or 'decked'"
    ),
    db: Session = Depends(get_db),
):
    """Get unique DB_ENTRY_TYPE values with counts"""
    query = db.query(MainDB.DB_ENTRY_TYPE, func.count(MainDB.DB_ID).label("count"))

    if status == "decked":
        _decked_ids = (
            db.query(ApplicationLogs.main_db_id)
            .filter(ApplicationLogs.application_step == "Decking")
            .subquery()
        )
        query = query.filter(
            or_(MainDB.DB_ID.in_(_decked_ids), MainDB.DB_APP_STATUS == "Completed")
        )
    elif status == "not_decked":
        _decked_ids = (
            db.query(ApplicationLogs.main_db_id)
            .filter(ApplicationLogs.application_step == "Decking")
            .subquery()
        )
        query = query.filter(
            MainDB.DB_ID.notin_(_decked_ids),
            or_(
                MainDB.DB_APP_STATUS.is_(None),
                MainDB.DB_APP_STATUS == "",
                MainDB.DB_APP_STATUS != "Completed",
            ),
        )

    results_with_type = (
        query.filter(MainDB.DB_ENTRY_TYPE.isnot(None), MainDB.DB_ENTRY_TYPE != "")
        .group_by(MainDB.DB_ENTRY_TYPE)
        .order_by(MainDB.DB_ENTRY_TYPE)
        .all()
    )

    query_no_type = db.query(func.count(MainDB.DB_ID))

    if status == "decked":
        _decked_ids = (
            db.query(ApplicationLogs.main_db_id)
            .filter(ApplicationLogs.application_step == "Decking")
            .subquery()
        )
        query_no_type = query_no_type.filter(
            or_(MainDB.DB_ID.in_(_decked_ids), MainDB.DB_APP_STATUS == "Completed")
        )
    elif status == "not_decked":
        _decked_ids = (
            db.query(ApplicationLogs.main_db_id)
            .filter(ApplicationLogs.application_step == "Decking")
            .subquery()
        )
        query_no_type = query_no_type.filter(
            MainDB.DB_ID.notin_(_decked_ids),
            or_(
                MainDB.DB_APP_STATUS.is_(None),
                MainDB.DB_APP_STATUS == "",
                MainDB.DB_APP_STATUS != "Completed",
            ),
        )

    no_type_count = query_no_type.filter(
        or_(MainDB.DB_ENTRY_TYPE.is_(None), MainDB.DB_ENTRY_TYPE == "")
    ).scalar()

    entry_types = [{"value": et, "count": count} for et, count in results_with_type]

    if no_type_count and no_type_count > 0:
        entry_types.insert(0, {"value": None, "count": no_type_count})

    return {"entry_types": entry_types}


@router.get("/logs/{main_id}", response_model=List[ApplicationLogResponse])
def get_logs(
    main_id: int,
    page: int = Query(1, ge=1),
    page_size: int = Query(10, ge=1, le=100),
    db: Session = Depends(get_db),
):
    """Get application logs for a specific MainDB record"""
    skip = (page - 1) * page_size
    logs, _ = get_application_logs(db=db, main_id=main_id, skip=skip, limit=page_size)
    return logs


@router.get("/summary", response_model=MainDBSummary)
def get_summary(db: Session = Depends(get_db)):
    """Get summary statistics"""
    return crud.get_main_db_summary(db)


@router.get("/filters/{field}")
def get_filter_options(field: str, db: Session = Depends(get_db)):
    """Get unique values for a field (for dropdown filters)"""
    values = crud.get_unique_values(db, field)
    return {"field": field, "values": values}


@router.post("/upload-excel")
async def upload_excel(
    file: UploadFile = File(...),
    username: str = Query("system"),
    db: Session = Depends(get_db),
):
    """Upload an Excel file and insert records into MainDB and ApplicationLogs"""
    print("🚀 Starting Excel upload process...")

    if not file.filename.endswith((".xls", ".xlsx")):
        raise HTTPException(
            status_code=400, detail="Invalid file type. Must be .xls or .xlsx"
        )

    try:
        contents = await file.read()
        df = pd.read_excel(io.BytesIO(contents))
    except Exception as e:
        raise HTTPException(
            status_code=400, detail=f"Failed to read Excel file: {str(e)}"
        )

    if df.empty:
        raise HTTPException(status_code=400, detail="Excel file is empty")

    for col in df.columns:
        if pd.api.types.is_datetime64_any_dtype(df[col]):
            df[col] = df[col].apply(
                lambda x: x.strftime("%Y-%m-%d") if pd.notna(x) else None
            )
        elif df[col].dtype == "object":
            if df[col].apply(lambda x: isinstance(x, pd.Timestamp)).any():
                df[col] = df[col].apply(
                    lambda x: (
                        x.strftime("%Y-%m-%d") if isinstance(x, pd.Timestamp) else x
                    )
                )

    print(f"📊 Total rows in Excel: {len(df)}")
    success, errors = 0, []
    duplicates_skipped = 0

    for index, row in df.iterrows():
        try:
            record_data = {}
            for excel_col, db_col in COLUMN_MAPPING.items():
                raw_value = row.get(excel_col)
                if pd.isna(raw_value) or raw_value is None:
                    record_data[db_col] = None
                elif isinstance(raw_value, (int, float, np.integer, np.floating)):
                    if db_col in NUMERIC_STRING_FIELDS:
                        record_data[db_col] = str(int(raw_value))
                    elif db_col in INTEGER_FIELDS:
                        record_data[db_col] = int(raw_value)
                    else:
                        record_data[db_col] = str(raw_value)
                else:
                    record_data[db_col] = (
                        str(raw_value).strip()
                        if isinstance(raw_value, str)
                        else str(raw_value)
                    )

            dtn_value = record_data.get("DB_DTN")
            if dtn_value:
                existing = crud.get_records_by_dtn(db, dtn_value)
                if existing:
                    duplicates_skipped += 1
                    errors.append(
                        {
                            "row_number": index + 2,
                            "dtn": str(dtn_value),
                            "brand_name": (
                                str(row.get("Brand Name", "-"))
                                if pd.notna(row.get("Brand Name", ""))
                                else "-"
                            ),
                            "reason": f"Duplicate DTN — already exists (DB_ID={[r.DB_ID for r in existing]})",
                        }
                    )
                    continue

            record_data["DB_USER_UPLOADER"] = username
            record_data["DB_DATE_EXCEL_UPLOAD"] = datetime.now().strftime(
                "%Y-%m-%d %H:%M:%S"
            )

            db_record = crud.create_main_db_record(
                db, MainDBCreate(**record_data), check_duplicate_dtn=False
            )
            print(
                f"  ✅ Inserted main_db ID {db_record.DB_ID} (DTN: {db_record.DB_DTN})"
            )

            logs_inserted = 0
            # ✅ FIXED — 8-value tuple, with proper id_col parsing
            for (
                step_label,
                user_col,
                id_col,
                decision_col,
                remarks_col,
                date_col,
                thread_col,
                del_idx,
            ) in LOG_STEPS:
                user_val = row.get(user_col)
                if pd.isna(user_val) or user_val is None or str(user_val).strip() == "":
                    continue

                # ✅ Parse numeric user_id
                raw_id = row.get(id_col)
                if raw_id is not None and not (
                    isinstance(raw_id, float) and pd.isna(raw_id)
                ):
                    try:
                        user_id_val = int(float(str(raw_id).strip()))
                    except (ValueError, TypeError):
                        user_id_val = None
                else:
                    user_id_val = None

                accomplished = parse_date_value(row.get(date_col))

                thread_val = row.get(thread_col)
                if (
                    not pd.isna(thread_val)
                    and thread_val is not None
                    and str(thread_val).strip() != ""
                ):
                    thread_str = str(thread_val).strip()
                else:
                    thread_str = "Open"

                # ✅ Mark COMPLETED if thread is "Close" OR if a date value is present
                has_date = accomplished is not None
                if thread_str.upper() == "CLOSE" or has_date:
                    thread_str = "Close" if thread_str.upper() == "CLOSE" else "Open"
                    del_last_index = 0
                    log_status = "COMPLETED"
                else:
                    thread_str = "Open"
                    del_last_index = 1
                    log_status = "IN PROGRESS"

                decision_val = row.get(decision_col)
                decision_str = (
                    str(decision_val).strip()
                    if not pd.isna(decision_val) and decision_val is not None
                    else ""
                )

                remarks_val = row.get(remarks_col)
                remarks_str = (
                    str(remarks_val).strip()
                    if not pd.isna(remarks_val) and remarks_val is not None
                    else ""
                )

                log = ApplicationLogs(
                    main_db_id=db_record.DB_ID,
                    application_step=step_label,
                    user_name=str(user_val).strip(),
                    user_id=user_id_val,  # ✅ Now properly defined
                    application_status=log_status,
                    application_decision=decision_str,
                    application_remarks=remarks_str,
                    start_date=accomplished,
                    accomplished_date=accomplished,
                    del_index=del_idx,
                    del_previous=None,
                    del_last_index=del_last_index,
                    del_thread=thread_str,
                )
                db.add(log)
                logs_inserted += 1
                print(
                    f"    📝 Log: {step_label} → {str(user_val).strip()} (ID={user_id_val}, del_index={del_idx})"
                )

            db.commit()
            print(
                f"  ✅ Committed {logs_inserted} log(s) for main_db ID {db_record.DB_ID}"
            )
            success += 1

        except Exception as e:
            db.rollback()
            print(f"❌ Error on row {index + 2}: {str(e)}")
            import traceback

            traceback.print_exc()
            errors.append(
                {
                    "row_number": index + 2,
                    "dtn": (
                        str(row.get("DTN", "-"))
                        if pd.notna(row.get("DTN", ""))
                        else "-"
                    ),
                    "brand_name": (
                        str(row.get("Brand Name", "-"))
                        if pd.notna(row.get("Brand Name", ""))
                        else "-"
                    ),
                    "reason": str(e),
                }
            )

    print(f"✅ Upload complete: {success} success, {len(errors)} errors")
    return {
        "success": True,
        "message": f"Upload complete: {success} records inserted successfully",
        "stats": {
            "total_processed": len(df),
            "success": success,
            "errors": len(errors),
            "duplicates_skipped": duplicates_skipped,
        },
        "failed_records": errors,
    }


@router.get("/download-template")
async def download_template():
    """Download Excel template with proper column headers"""
    try:
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter

        log_step_columns = []
        for (
            step_label,
            user_col,
            id_col,
            decision_col,
            remarks_col,
            date_col,
            thread_col,
            _,
        ) in LOG_STEPS:
            log_step_columns += [
                user_col,
                id_col,
                decision_col,
                remarks_col,
                date_col,
                thread_col,
            ]

        all_columns = list(COLUMN_MAPPING.keys()) + log_step_columns
        template_data = {col: [""] for col in all_columns}
        df = pd.DataFrame(template_data)

        # Color scheme per step
        STEP_COLORS = [
            "FFF2CC",  # Decker          — yellow
            "D9EAD3",  # Evaluator        — green
            "CFE2F3",  # Checker          — blue
            "EAD1DC",  # Supervisor       — pink
            "D9D2E9",  # QA Admin         — purple
            "FCE5CD",  # LRD Chief Admin  — orange
            "D0E4F7",  # OD-Receiving     — light blue
            "F4CCCC",  # OD-Releasing     — red/rose
            "D9EAD3",  # Releasing Officer— green (different shade optional)
        ]

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Template")

            workbook = writer.book
            worksheet = writer.sheets["Template"]

            main_col_count = len(COLUMN_MAPPING)

            # Style the main DB columns header — subtle gray
            main_fill = PatternFill(
                start_color="EFEFEF", end_color="EFEFEF", fill_type="solid"
            )
            main_font = Font(bold=True)
            for col_idx in range(1, main_col_count + 1):
                cell = worksheet.cell(row=1, column=col_idx)
                cell.fill = main_fill
                cell.font = main_font
                cell.alignment = Alignment(horizontal="center", wrap_text=True)
                worksheet.column_dimensions[get_column_letter(col_idx)].width = 18

            # Style the log step columns — colored per step (6 cols each)
            log_col_start = main_col_count + 1
            for step_idx, step_tuple in enumerate(LOG_STEPS):
                color = STEP_COLORS[step_idx % len(STEP_COLORS)]
                fill = PatternFill(
                    start_color=color, end_color=color, fill_type="solid"
                )
                bold_font = Font(bold=True)

                for offset in range(
                    6
                ):  # 6 cols per step: name, id, decision, remarks, date, thread
                    col_idx = log_col_start + (step_idx * 6) + offset
                    cell = worksheet.cell(row=1, column=col_idx)
                    cell.fill = fill
                    cell.font = bold_font
                    cell.alignment = Alignment(horizontal="center", wrap_text=True)
                    worksheet.column_dimensions[get_column_letter(col_idx)].width = 20

            # Freeze the first row
            worksheet.freeze_panes = "A2"

            # Instructions sheet
            instructions = pd.DataFrame(
                {
                    "Column Group": [
                        "Main Database Columns",
                        "Application Log Columns",
                        "ID Columns",
                        "Del Thread Values",
                        "Date Format Instructions",
                        "Numeric Field Instructions",
                    ],
                    "Description": [
                        "Columns from 'DTN' to 'Processing Type' — gray highlighted",
                        "Each step has 6 columns: Name, ID, Decision, Remarks, Date, Del Thread — color coded per step",
                        "ID columns (e.g. 'Decker ID', 'QA Admin ID') — numeric employee ID (e.g. 1001)",
                        "Del Thread column per step — values: 'Open' or 'Close'",
                        "For date fields, use formats like: 2026-01-02, Jan 2 2026, 01/02/2026",
                        "Timeline Citizen Charter should be a whole number (e.g., 30, 45, 60)",
                    ],
                    "Note": [
                        "All main database columns are optional",
                        "A log row is only inserted if the Name column (e.g., 'Decker') has a value",
                        "Leave blank if no employee ID. Must be a whole number if filled.",
                        "Leave blank if not applicable — defaults to 'Open'",
                        "Date fields will be automatically parsed. Leave empty if no date.",
                        "Enter numbers without decimals for timeline fields.",
                    ],
                }
            )
            instructions.to_excel(writer, index=False, sheet_name="Instructions")

            # Style instructions sheet header too
            inst_ws = writer.sheets["Instructions"]
            inst_fill = PatternFill(
                start_color="4A90D9", end_color="4A90D9", fill_type="solid"
            )
            inst_font = Font(bold=True, color="FFFFFF")
            for col_idx in range(1, 4):
                cell = inst_ws.cell(row=1, column=col_idx)
                cell.fill = inst_fill
                cell.font = inst_font
                cell.alignment = Alignment(horizontal="center")
                inst_ws.column_dimensions[get_column_letter(col_idx)].width = 35

        output.seek(0)
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": "attachment; filename=main_db_template.xlsx"
            },
        )
    except Exception as e:
        raise HTTPException(
            status_code=500, detail=f"Failed to generate template: {str(e)}"
        )


@router.get("/upload-history")
async def get_upload_history(
    limit: int = Query(10, ge=1, le=200),
    offset: int = Query(0, ge=0),
    impersonated_user_id: Optional[int] = Query(None),
    date_from: Optional[date] = Query(None),
    date_to: Optional[date] = Query(None),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_active_user),
):
    """Get upload history grouped by upload batch (per user), with optional date filtering."""
    try:
        target_username = current_user.username

        if impersonated_user_id and current_user.role.value in ("Admin", "SuperAdmin"):
            target = db.query(User).filter(User.id == impersonated_user_id).first()
            if target:
                target_username = target.username

        data, total, total_records = get_upload_history_paginated(
            db=db,
            username=target_username,
            limit=limit,
            offset=offset,
            date_from=date_from,
            date_to=date_to,
        )

        return {
            "success": True,
            "total": total,
            "total_records": total_records,
            "data": data,
        }

    except Exception as e:
        raise HTTPException(
            status_code=500, detail=f"Failed to fetch upload history: {str(e)}"
        )


@router.get("/export-filtered")
async def export_filtered_records(
    search: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    category: Optional[str] = Query(None),
    prescription: Optional[str] = Query(None),
    prescription_not: Optional[str] = Query(None),
    dtn: Optional[int] = Query(None),
    manufacturer: Optional[str] = Query(None),
    lto_company: Optional[str] = Query(None),
    brand_name: Optional[str] = Query(None),
    generic_name: Optional[str] = Query(None),
    app_status: Optional[str] = Query(None),
    app_type: Optional[str] = Query(None),
    processing_type: Optional[str] = Query(None),
    entry_type: Optional[str] = Query(None),
    dosage_form: Optional[str] = Query(None),
    manufacturer_country: Optional[str] = Query(None),
    trader: Optional[str] = Query(None),
    trader_country: Optional[str] = Query(None),
    importer: Optional[str] = Query(None),
    importer_country: Optional[str] = Query(None),
    distributor: Optional[str] = Query(None),
    distributor_country: Optional[str] = Query(None),
    repacker: Optional[str] = Query(None),
    repacker_country: Optional[str] = Query(None),
    type_doc_released: Optional[str] = Query(None),
    date_released_from: Optional[str] = Query(None),
    date_released_to: Optional[str] = Query(None),
    date_received_cent_from: Optional[str] = Query(None),
    date_received_cent_to: Optional[str] = Query(None),
    null_date_released: Optional[str] = Query(None),
    null_date_received_cent: Optional[str] = Query(None),
    columns: Optional[str] = Query(
        None,
        description="Comma-separated column ids to include. Omit = all columns.",
    ),
    db: Session = Depends(get_db),
):
    """Export filtered records to Excel — streamed, only selected columns"""
    try:
        print(
            f"📥 Export request: status={status}, app_type={app_type}, search={search}, columns={columns}"
        )

        # ── Parse selected columns ──
        selected_ids = None
        if columns:
            selected_ids = {c.strip() for c in columns.split(",") if c.strip()}

        selected_main_cols = [
            c for c in EXPORT_COLUMNS if selected_ids is None or c[0] in selected_ids
        ]

        all_log_ids = {step[0]: _log_step_export_id(step[0]) for step in LOG_STEPS}
        selected_log_steps = (
            LOG_STEPS
            if selected_ids is None
            else [s for s in LOG_STEPS if all_log_ids[s[0]] in selected_ids]
        )
        include_entry_type = selected_ids is None or "entry_type" in selected_ids

        filters = {
            "status": status,
            "category": category,
            "prescription": prescription,
            "prescription_not": prescription_not,
            "dtn": dtn,
            "manufacturer": manufacturer,
            "lto_company": lto_company,
            "brand_name": brand_name,
            "generic_name": generic_name,
            "app_status": app_status,
            "app_type": app_type,
            "processing_type": processing_type,
            "entry_type": entry_type,
            "dosage_form": dosage_form,
            "manufacturer_country": manufacturer_country,
            "trader": trader,
            "trader_country": trader_country,
            "importer": importer,
            "importer_country": importer_country,
            "distributor": distributor,
            "distributor_country": distributor_country,
            "repacker": repacker,
            "repacker_country": repacker_country,
            "type_doc_released": type_doc_released,
            "date_released_from": date_released_from,
            "date_released_to": date_released_to,
            "date_received_cent_from": date_received_cent_from,
            "date_received_cent_to": date_received_cent_to,
            "null_date_released": null_date_released,
            "null_date_received_cent": null_date_received_cent,
        }

        records, total = get_main_db_records(
            db=db,
            skip=0,
            limit=100000,
            search=search,
            filters=filters,
            sort_by="DB_DATE_EXCEL_UPLOAD",
            sort_order="desc",
        )

        print(f"📊 Exporting {total} records with {len(selected_main_cols)} columns")

        if not records:
            raise HTTPException(status_code=404, detail="No records found to export")

        from sqlalchemy.orm import joinedload

        main_db_ids = [r.DB_ID for r in records]

        latest_by_step = {}
        if selected_log_steps:
            step_labels = [s[0] for s in selected_log_steps]
            logs = (
                db.query(ApplicationLogs)
                .options(joinedload(ApplicationLogs.user))
                .filter(ApplicationLogs.main_db_id.in_(main_db_ids))
                .filter(ApplicationLogs.application_step.in_(step_labels))
                .order_by(ApplicationLogs.main_db_id, ApplicationLogs.created_at.desc())
                .all()
            )

            def _log_user_name(log):
                if log.user:
                    name = (
                        f"{log.user.first_name or ''} {log.user.surname or ''}".strip()
                    )
                    if name:
                        return name
                return log.user_name

            for log in logs:
                bucket = latest_by_step.setdefault(log.main_db_id, {})
                if log.application_step not in bucket:
                    bucket[log.application_step] = log
        else:

            def _log_user_name(log):
                return None

        LOG_STEP_COLORS = [
            "FFF2CC",
            "D9EAD3",
            "FDE2C8",
            "CFE2F3",
            "EAD1DC",
            "D9D2E9",
            "FCE5CD",
            "D0E4F7",
            "F4CCCC",
            "C9DAF8",
        ]

        from openpyxl import Workbook
        from openpyxl.cell import WriteOnlyCell
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        wb = Workbook(write_only=True)
        ws = wb.create_sheet("Filtered Records")

        HEADERS = [label for _, label, _, _ in selected_main_cols]
        WIDTHS = [width for _, _, _, width in selected_main_cols]

        LOG_COLUMN_FILL_BY_INDEX = {}
        for (
            step_label,
            user_col,
            id_col,
            decision_col,
            remarks_col,
            date_col,
            thread_col,
            del_idx,
        ) in selected_log_steps:
            step_global_idx = next(
                i for i, s in enumerate(LOG_STEPS) if s[0] == step_label
            )
            color = LOG_STEP_COLORS[step_global_idx % len(LOG_STEP_COLORS)]
            start_col = len(HEADERS) + 1
            HEADERS += [user_col, decision_col, remarks_col, date_col]
            WIDTHS += [25, 22, 30, 20]
            for offset in range(4):
                LOG_COLUMN_FILL_BY_INDEX[start_col + offset] = color

        if include_entry_type:
            HEADERS.append("Entry Type")
            WIDTHS.append(20)

        for col_idx, width in enumerate(WIDTHS, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        header_fill = PatternFill(
            start_color="1F4E79", end_color="1F4E79", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF", size=10)
        header_align = Alignment(horizontal="center", vertical="center")
        log_header_font = Font(bold=True, color="000000", size=10)

        header_cells = []
        for col_idx, h in enumerate(HEADERS, start=1):
            c = WriteOnlyCell(ws, value=h)
            if col_idx in LOG_COLUMN_FILL_BY_INDEX:
                c.fill = PatternFill(
                    start_color=LOG_COLUMN_FILL_BY_INDEX[col_idx],
                    end_color=LOG_COLUMN_FILL_BY_INDEX[col_idx],
                    fill_type="solid",
                )
                c.font = log_header_font
            else:
                c.fill = header_fill
                c.font = header_font
            c.alignment = header_align
            header_cells.append(c)
        ws.append(header_cells)

        even_fill = PatternFill(
            start_color="EBF3FB", end_color="EBF3FB", fill_type="solid"
        )
        data_align = Alignment(vertical="center")
        dtn_col_idx = next(
            (i + 1 for i, (cid, *_r) in enumerate(selected_main_cols) if cid == "dtn"),
            None,
        )

        for row_idx, record in enumerate(records, start=2):
            step_logs = latest_by_step.get(record.DB_ID, {})
            values = [getter(record) for _, _, getter, _ in selected_main_cols]

            for step_label, *_rest in selected_log_steps:
                log = step_logs.get(step_label)
                values += [
                    _log_user_name(log) if log else None,
                    log.application_decision if log else None,
                    log.application_remarks if log else None,
                    (
                        str(log.accomplished_date)
                        if log and log.accomplished_date
                        else None
                    ),
                ]

            if include_entry_type:
                values.append(record.DB_ENTRY_TYPE)

            is_even = row_idx % 2 == 0
            row_cells = []
            for col_idx, val in enumerate(values, start=1):
                c = WriteOnlyCell(ws, value=val)
                c.alignment = data_align
                if col_idx in LOG_COLUMN_FILL_BY_INDEX:
                    c.fill = PatternFill(
                        start_color=LOG_COLUMN_FILL_BY_INDEX[col_idx],
                        end_color=LOG_COLUMN_FILL_BY_INDEX[col_idx],
                        fill_type="solid",
                    )
                elif is_even:
                    c.fill = even_fill
                if dtn_col_idx and col_idx == dtn_col_idx and val is not None:
                    c.number_format = "@"
                row_cells.append(c)
            ws.append(row_cells)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        parts = ["main_db_export"]
        if app_type:
            parts.append(app_type.replace(" ", "_"))
        if status:
            parts.append(status)
        parts.append(timestamp)
        filename = "_".join(parts) + ".xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}",
                "Content-Length": str(output.getbuffer().nbytes),
            },
        )

    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Export error: {str(e)}")
        import traceback

        traceback.print_exc()
        raise HTTPException(
            status_code=500, detail=f"Failed to export records: {str(e)}"
        )


@router.get("/export-columns")
def get_export_columns():
    """List of columns available for export selection"""
    cols = [
        {"id": cid, "label": label, "group": "Main"}
        for cid, label, _, _ in EXPORT_COLUMNS
    ]
    cols += [
        {
            "id": _log_step_export_id(step[0]),
            "label": step[0],
            "group": "Application Log",
        }
        for step in LOG_STEPS
    ]
    cols.append({"id": "entry_type", "label": "Entry Type", "group": "Main"})
    return {"columns": cols}


@router.get("/{record_id}", response_model=MainDBResponse)
def get_record(record_id: int, db: Session = Depends(get_db)):
    """Get a single record by ID"""
    record = crud.get_main_db_record(db, record_id)
    if not record:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Record with ID {record_id} not found",
        )
    return record


@router.post("/", response_model=MainDBResponse, status_code=status.HTTP_201_CREATED)
def create_record(record: MainDBCreate, db: Session = Depends(get_db)):
    """Create a new record"""
    return crud.create_main_db_record(db, record)


@router.post(
    "/bulk", response_model=List[MainDBResponse], status_code=status.HTTP_201_CREATED
)
def create_bulk_records(records: List[MainDBCreate], db: Session = Depends(get_db)):
    """Bulk create records"""
    return crud.bulk_create_main_db_records(db, records)


@router.put("/{record_id}", response_model=MainDBResponse)
def update_record(
    record_id: int, record_update: MainDBUpdate, db: Session = Depends(get_db)
):
    """Update an existing record"""
    updated_record = crud.update_main_db_record(db, record_id, record_update)
    if not updated_record:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Record with ID {record_id} not found",
        )
    return updated_record


@router.delete("/{record_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_record(
    record_id: int,
    hard_delete: bool = Query(
        False, description="Permanently delete (default: soft delete)"
    ),
    db: Session = Depends(get_db),
):
    """Delete a record (soft delete by default)"""
    if hard_delete:
        success = crud.hard_delete_main_db_record(db, record_id)
    else:
        success = crud.delete_main_db_record(db, record_id)

    if not success:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Record with ID {record_id} not found",
        )
    return None


@router.post("/{record_id}/restore", response_model=MainDBResponse)
def restore_record(record_id: int, db: Session = Depends(get_db)):
    """Restore a soft-deleted record"""
    record = crud.get_main_db_record(db, record_id)
    if not record:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Record with ID {record_id} not found",
        )
    record.DB_TRASH = None
    record.DB_TRASH_DATE_ENCODED = None
    db.commit()
    db.refresh(record)
    return record
