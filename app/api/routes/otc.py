from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import Optional
import pandas as pd
import io
import numpy as np
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.db.remote_session import RemoteOTCSessionLocal
from app.schemas.otc import OTCUploadResponse, OTCRecordUpdate
from app.crud import otc as crud
from app.core.deps import get_current_active_user

router = APIRouter(
    prefix="/api/otc",
    tags=["OTC Database"],
    dependencies=[Depends(get_current_active_user)]
)


# ---------------------
# DB Dependency
# ---------------------
def get_otc_db():
    if RemoteOTCSessionLocal is None:
        raise HTTPException(
            status_code=503,
            detail="OTC database connection is not configured."
        )
    db = RemoteOTCSessionLocal()
    try:
        yield db
    finally:
        db.close()


# ---------------------
# Column Mapping (human-readable Excel header → DB column)
# ---------------------
COLUMN_MAPPING = {
    "DTN": "DB_DTN",
    "Est. Category": "DB_EST_CAT",
    "LTO Company": "DB_EST_LTO_COMP",
    "LTO Address": "DB_EST_LTO_ADD",
    "Email": "DB_EST_EADD",
    "TIN": "DB_EST_TIN",
    "Contact No.": "DB_EST_CONTACT_NO",
    "LTO No.": "DB_EST_LTO_NO",
    "Validity": "DB_EST_VALIDITY",
    "Brand Name": "DB_PROD_BR_NAME",
    "Generic Name": "DB_PROD_GEN_NAME",
    "Dosage Strength": "DB_PROD_DOS_STR",
    "Dosage Form": "DB_PROD_DOS_FORM",
    "Prescription": "DB_PROD_CLASS_PRESCRIP",
    "Essential Drug": "DB_PROD_ESS_DRUG_LIST",
    "Pharma Category": "DB_PROD_PHARMA_CAT",
    "Manufacturer": "DB_PROD_MANU",
    "Manufacturer Address": "DB_PROD_MANU_ADD",
    "Manufacturer TIN": "DB_PROD_MANU_TIN",
    "Manufacturer LTO No.": "DB_PROD_MANU_LTO_NO",
    "Manufacturer Country": "DB_PROD_MANU_COUNTRY",
    "Trader": "DB_PROD_TRADER",
    "Trader Address": "DB_PROD_TRADER_ADD",
    "Trader TIN": "DB_PROD_TRADER_TIN",
    "Trader LTO No.": "DB_PROD_TRADER_LTO_NO",
    "Trader Country": "DB_PROD_TRADER_COUNTRY",
    "Repacker": "DB_PROD_REPACKER",
    "Repacker Address": "DB_PROD_REPACKER_ADD",
    "Repacker TIN": "DB_PROD_REPACKER_TIN",
    "Repacker LTO No.": "DB_PROD_REPACKER_LTO_NO",
    "Repacker Country": "DB_PROD_REPACKER_COUNTRY",
    "Importer": "DB_PROD_IMPORTER",
    "Importer Address": "DB_PROD_IMPORTER_ADD",
    "Importer TIN": "DB_PROD_IMPORTER_TIN",
    "Importer LTO No.": "DB_PROD_IMPORTER_LTO_NO",
    "Importer Country": "DB_PROD_IMPORTER_COUNTRY",
    "Distributor": "DB_PROD_DISTRI",
    "Distributor Address": "DB_PROD_DISTRI_ADD",
    "Distributor TIN": "DB_PROD_DISTRI_TIN",
    "Distributor LTO No.": "DB_PROD_DISTRI_LTO_NO",
    "Distributor Country": "DB_PROD_DISTRI_COUNTRY",
    "Shelf Life": "DB_PROD_DISTRI_SHELF_LIFE",
    "Storage Condition": "DB_STORAGE_COND",
    "Packaging": "DB_PACKAGING",
    "Suggested RP": "DB_SUGG_RP",
    "No. Sample": "DB_NO_SAMPLE",
    "Expiry Date": "DB_EXPIRY_DATE",
    "CPR Validity": "DB_CPR_VALIDITY",
    "Registration No.": "DB_REG_NO",
    "App Type": "DB_APP_TYPE",
    "Mother App Type": "DB_MOTHER_APP_TYPE",
    "Old RSN": "DB_OLD_RSN",
    "Amendment 1": "DB_AMMEND1",
    "Amendment 2": "DB_AMMEND2",
    "Amendment 3": "DB_AMMEND3",
    "Product Category": "DB_PROD_CAT",
    "Certification": "DB_CERTIFICATION",
    "Fee": "DB_FEE",
    "LRF": "DB_LRF",
    "SURC": "DB_SURC",
    "Total": "DB_TOTAL",
    "OR No.": "DB_OR_NO",
    "Date Issued": "DB_DATE_ISSUED",
    "Date Received FDAC": "DB_DATE_RECEIVED_FDAC",
    "Date Received Central": "DB_DATE_RECEIVED_CENT",
    "MO": "DB_MO",
    "File": "DB_FILE",
    "SECPA": "DB_SECPA",
    "SECPA Exp Date": "DB_SECPA_EXP_DATE",
    "SECPA Issued On": "DB_SECPA_ISSUED_ON",
    "Decking Schedule": "DB_DECKING_SCHED",
    "Evaluation": "DB_EVAL",
    "Date Deck": "DB_DATE_DECK",
    "Remarks 1": "DB_REMARKS_1",
    "Date Remarks": "DB_DATE_REMARKS",
    "Class": "DB_CLASS",
    "Date Released": "DB_DATE_RELEASED",
    "Type Doc Released": "DB_TYPE_DOC_RELEASED",
    "Atta Released": "DB_ATTA_RELEASED",
    "CPR Condition": "DB_CPR_COND",
    "CPR Cond Remarks": "DB_CPR_COND_REMARKS",
    "CPR Cond Add Remarks": "DB_CPR_COND_ADD_REMARKS",
    "App Status": "DB_APP_STATUS",
    "Pharma Prod Cat": "DB_PHARMA_PROD_CAT",
    "Pharma Prod Cat Label": "DB_PHARMA_PROD_CAT_LABEL",
    "Is in PM": "DB_IS_IN_PM",
    "Timeline Citizen Charter": "DB_TIMELINE_CITIZEN_CHARTER",
}

NUMERIC_STRING_FIELDS = {"DB_FEE", "DB_LRF", "DB_SURC", "DB_TOTAL"}
INTEGER_FIELDS = {"DB_DTN", "DB_IS_IN_PM", "DB_TIMELINE_CITIZEN_CHARTER"}
DATE_FIELDS = {"DB_DATE_ISSUED"}


# ---------------------
# Helper
# ---------------------
def clean_value(raw, db_col):
    """Normalize a raw Excel cell value to the correct Python type for the DB column."""
    if raw is None or (isinstance(raw, float) and np.isnan(raw)):
        return None

    if db_col in INTEGER_FIELDS:
        try:
            return int(raw)
        except (ValueError, TypeError):
            return None

    if db_col in NUMERIC_STRING_FIELDS:
        try:
            return str(int(float(raw)))
        except (ValueError, TypeError):
            return str(raw).strip()

    if db_col in DATE_FIELDS:
        if isinstance(raw, (datetime, pd.Timestamp)):
            return raw.date()
        if isinstance(raw, str):
            raw = raw.strip()
            if not raw:
                return None
            try:
                from dateutil import parser as dateparser
                return dateparser.parse(raw, fuzzy=True).date()
            except Exception:
                return None
        return None

    # Default: string
    if isinstance(raw, (int, float, np.integer, np.floating)):
        return str(raw)
    return str(raw).strip() if isinstance(raw, str) else str(raw)


# ---------------------
# Routes
# ---------------------

@router.get("/download-template")
async def download_otc_template():
    """Download a formatted OTC Excel upload template."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "OTC Upload Template"

    headers = list(COLUMN_MAPPING.keys())

    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill("solid", start_color="1F4E79")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin
        ws.column_dimensions[get_column_letter(col_idx)].width = 22

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 35

    # Instructions sheet
    ws_info = wb.create_sheet("Instructions")
    ws_info["A1"] = "OTC UPLOAD TEMPLATE - INSTRUCTIONS"
    ws_info["A1"].font = Font(name="Arial", bold=True, size=14, color="1F4E79")

    instruction_rows = [
        ("Column", "Data Type", "Notes"),
        ("DTN", "Number (bigint)", "Document Tracking Number — numeric only"),
        ("Date Issued", "Date", "Format: YYYY-MM-DD or MM/DD/YYYY"),
        ("App Status", "Text (max 50 chars)", "Application Status"),
        ("Is in PM", "Number (0 or 1)", "1 = Yes, 0 = No"),
        ("Timeline Citizen Charter", "Integer", "Whole number only (e.g. 30, 45, 60)"),
        ("Fee / LRF / SURC / Total", "Number", "Stored as numeric string"),
    ]

    for r_idx, (col, dtype, note) in enumerate(instruction_rows, start=3):
        ws_info.cell(row=r_idx, column=1, value=col)
        ws_info.cell(row=r_idx, column=2, value=dtype)
        ws_info.cell(row=r_idx, column=3, value=note)
        if r_idx == 3:
            for c in range(1, 4):
                cell = ws_info.cell(row=r_idx, column=c)
                cell.font = Font(name="Arial", bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", start_color="1F4E79")

    ws_info.column_dimensions["A"].width = 35
    ws_info.column_dimensions["B"].width = 25
    ws_info.column_dimensions["C"].width = 45

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=otc_upload_template.xlsx"},
    )


@router.post("/upload-excel", response_model=OTCUploadResponse)
async def upload_otc_excel(
    file: UploadFile = File(...),
    username: str = Query("system"),
    db: Session = Depends(get_otc_db),
):
    """Upload an Excel file and insert OTC records into the remote wf_cdrr database."""
    if not file.filename.endswith((".xls", ".xlsx")):
        raise HTTPException(status_code=400, detail="Invalid file type. Must be .xls or .xlsx")

    try:
        contents = await file.read()
        df = pd.read_excel(io.BytesIO(contents))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to read Excel file: {str(e)}")

    if df.empty:
        raise HTTPException(status_code=400, detail="Excel file is empty")

    print(f"📊 OTC Upload — total rows: {len(df)}")
    success, errors = 0, []

    for index, row in df.iterrows():
        try:
            record_data = {}

            for excel_col, db_col in COLUMN_MAPPING.items():
                raw = row.get(excel_col)
                record_data[db_col] = clean_value(raw, db_col)

            record_data["DB_USER_UPLOADER"] = username
            record_data["DB_DATE_EXCEL_UPLOAD"] = datetime.now()

            # Lagyan ng default value kung walang laman
            if not record_data.get("DB_APP_STATUS"):
                record_data["DB_APP_STATUS"] = "TO_DO"

            db_record = crud.create_otc_record(db, record_data)
            success += 1
            print(f"  ✅ Row {index + 2} → ID {db_record['DB_ID']}")

        except Exception as e:
            print(f"  ❌ Row {index + 2}: {e}")
            errors.append({
                "row": index + 2,
                "error": str(e),
                "data": {
                    k: str(v)[:50]
                    for k, v in row.to_dict().items()
                    if not (isinstance(v, float) and np.isnan(v))
                },
            })

    print(f"✅ OTC Upload complete — success: {success}, errors: {len(errors)}")

    return OTCUploadResponse(
        success=True,
        message=f"Upload complete: {success} records inserted successfully",
        stats={"total": len(df), "success": success, "errors": len(errors)},
        errors=errors[:10],
    )


@router.get("/records")
async def get_otc_records_list(
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=500),
    search: Optional[str] = Query(None),
    app_status: Optional[str] = Query(None),
    prescription: Optional[str] = Query(None),
    brand_name: Optional[str] = Query(None),
    generic_name: Optional[str] = Query(None),
    lto_company: Optional[str] = Query(None),
    registration_no: Optional[str] = Query(None),
    app_type: Optional[str] = Query(None),
    is_in_pm: Optional[str] = Query(None),  # ✅ CHANGED: from decking_status to is_in_pm
    sort_by: Optional[str] = Query("DB_DATE_EXCEL_UPLOAD"),
    sort_order: Optional[str] = Query("desc"),
    db: Session = Depends(get_otc_db),
):
    """
    Get OTC records with comprehensive pagination and filters.
    
    - **skip**: Number of records to skip (default: 0)
    - **limit**: Number of records to return (default: 100, max: 500)
    - **search**: General search across brand name, generic name, registration no, LTO company, DTN
    - **app_status**: Filter by application status
    - **prescription**: Filter by prescription type
    - **brand_name**: Filter by brand name (partial match)
    - **generic_name**: Filter by generic name (partial match)
    - **lto_company**: Filter by LTO company (partial match)
    - **registration_no**: Filter by registration number (partial match)
    - **app_type**: Filter by application type
    - **is_in_pm**: Filter by PM status - 'not_in_pm' (blank/N/A) or 'in_pm' (value=1)
    - **sort_by**: Column to sort by (default: DB_DATE_EXCEL_UPLOAD)
    - **sort_order**: Sort order - 'asc' or 'desc' (default: desc)
    """
    result = crud.get_otc_records(
        db=db,
        skip=skip,
        limit=limit,
        search=search,
        app_status=app_status,
        prescription=prescription,
        brand_name=brand_name,
        generic_name=generic_name,
        lto_company=lto_company,
        registration_no=registration_no,
        app_type=app_type,
        is_in_pm=is_in_pm,  # ✅ CHANGED: pass is_in_pm instead of decking_status
        sort_by=sort_by,
        sort_order=sort_order
    )
    
    return result


@router.get("/records/{record_id}")
async def get_otc_record_detail(
    record_id: int,
    db: Session = Depends(get_otc_db),
):
    """Get a single OTC record by ID."""
    record = crud.get_otc_record_by_id(db, record_id)
    
    if not record:
        raise HTTPException(status_code=404, detail=f"Record with ID {record_id} not found")
    
    return record


@router.put("/records/{record_id}")
async def update_otc_record(
    record_id: int,
    update_data: dict,
    db: Session = Depends(get_otc_db),
):
    """
    Update an OTC record by ID.
    
    - **record_id**: The ID of the record to update
    - **update_data**: Dictionary of fields to update
    """
    # Validate that the record exists
    existing = crud.get_otc_record_by_id(db, record_id)
    if not existing:
        raise HTTPException(status_code=404, detail=f"Record with ID {record_id} not found")
    
    # Update the record
    updated = crud.update_otc_record(db, record_id, update_data)
    
    if not updated:
        raise HTTPException(status_code=500, detail="Failed to update record")
    
    return {
        "success": True,
        "message": "Record updated successfully",
        "record": updated
    }


@router.delete("/records/{record_id}")
async def delete_otc_record(
    record_id: int,
    db: Session = Depends(get_otc_db),
):
    """
    Delete an OTC record by ID.
    
    - **record_id**: The ID of the record to delete
    """
    # Check if record exists
    existing = crud.get_otc_record_by_id(db, record_id)
    if not existing:
        raise HTTPException(status_code=404, detail=f"Record with ID {record_id} not found")
    
    # Delete the record
    deleted = crud.delete_otc_record(db, record_id)
    
    if not deleted:
        raise HTTPException(status_code=500, detail="Failed to delete record")
    
    return {
        "success": True,
        "message": f"Record with ID {record_id} deleted successfully"
    }


@router.get("/app-statuses")
async def get_app_statuses(
    is_in_pm: Optional[str] = Query(None),
    app_type: Optional[str] = Query(None),
    prescription: Optional[str] = Query(None),
    app_status: Optional[str] = Query(None),
    db: Session = Depends(get_otc_db),
):
    """Get unique application statuses with counts for filter options."""
    statuses = crud.get_app_statuses(
        db,
        is_in_pm=is_in_pm,
        app_type=app_type,
        prescription=prescription,
    )
    return {"app_statuses": statuses}


@router.get("/app-types")
async def get_app_types(
    is_in_pm: Optional[str] = Query(None),
    prescription: Optional[str] = Query(None),
    app_status: Optional[str] = Query(None),
    db: Session = Depends(get_otc_db),
):
    """Get unique application types with counts for filter options."""
    types = crud.get_app_types(
        db,
        is_in_pm=is_in_pm,
        prescription=prescription,
        app_status=app_status,
    )
    return {"app_types": types}


@router.get("/prescription-types")
async def get_prescription_types(
    is_in_pm: Optional[str] = Query(None),
    app_type: Optional[str] = Query(None),
    app_status: Optional[str] = Query(None),
    db: Session = Depends(get_otc_db),
):
    """Get unique prescription types with counts for filter options."""
    types = crud.get_prescription_types(
        db,
        is_in_pm=is_in_pm,
        app_type=app_type,
        app_status=app_status,
    )
    return {"prescription_types": types}


@router.get("/export-filtered")
async def export_filtered_records(
    search: Optional[str] = Query(None),
    app_status: Optional[str] = Query(None),
    prescription: Optional[str] = Query(None),
    brand_name: Optional[str] = Query(None),
    generic_name: Optional[str] = Query(None),
    lto_company: Optional[str] = Query(None),
    registration_no: Optional[str] = Query(None),
    app_type: Optional[str] = Query(None),
    is_in_pm: Optional[str] = Query(None),  # ✅ CHANGED: from decking_status to is_in_pm
    db: Session = Depends(get_otc_db),
):
    """
    Export filtered OTC records to Excel.
    
    All filter parameters are optional and work the same as the /records endpoint.
    Returns an Excel file with all matching records (no pagination).
    """
    from openpyxl import Workbook
    
    # Get filtered records
    records = crud.export_otc_records_data(
        db=db,
        search=search,
        app_status=app_status,
        prescription=prescription,
        brand_name=brand_name,
        generic_name=generic_name,
        lto_company=lto_company,
        registration_no=registration_no,
        app_type=app_type,
        is_in_pm=is_in_pm  # ✅ CHANGED: pass is_in_pm instead of decking_status
    )
    
    if not records:
        raise HTTPException(status_code=404, detail="No records found matching the filter criteria")
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "OTC Export"
    
    # Headers (reverse mapping from DB columns to human-readable)
    db_to_excel = {v: k for k, v in COLUMN_MAPPING.items()}
    
    # Get all columns from the first record
    all_columns = list(records[0].keys())
    
    # Header styling
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill("solid", start_color="1F4E79")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Write headers
    for col_idx, db_col in enumerate(all_columns, start=1):
        excel_col = db_to_excel.get(db_col, db_col)  # Use DB name if no mapping exists
        cell = ws.cell(row=1, column=col_idx, value=excel_col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        ws.column_dimensions[get_column_letter(col_idx)].width = 20
    
    # Write data
    for row_idx, record in enumerate(records, start=2):
        for col_idx, db_col in enumerate(all_columns, start=1):
            value = record.get(db_col)
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    ws.freeze_panes = "A2"
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Generate filename with timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"otc_export_{timestamp}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


