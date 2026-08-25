# app/crud/gmp_upload.py
"""
GMP Excel template builder and upload parser.
Mirrors the exact pattern used in main_db.py:
  - pandas + openpyxl for template generation
  - pandas for upload parsing
  - Same color-per-step-group header styling
  - Instructions sheet included
"""

import io
import pandas as pd
import numpy as np
from datetime import datetime
from dateutil import parser as dateutil_parser
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# Column mapping — Excel header → GMPRecord field
# ─────────────────────────────────────────────────────────────────────────────
GMP_COLUMN_MAPPING = {
    "DTN":                                          "GMP_DTN",
    "RELATED DTN":                                  "GMP_RELATED_DTN",
    "DATE RECEIVED":                                "GMP_DATE_RECEIVED",
    "NAME OF ESTABLISHMENT":                        "GMP_LTO_COMPANY",
    "LTO NUMBER":                                   "GMP_LTO_NUMBER",
    "ADDRESS":                                      "GMP_LTO_ADDRESS",
    "TRANSACTION TYPE":                             "GMP_TRANSACTION_TYPE",
    "CATEGORY":                                     "GMP_EST_CATEGORY",
    "FOREIGN MANUFACTURER":                         "GMP_FOREIGN_MANUFACTURER",
    "FOREIGN MANUFACTURER ADDRESS":                 "GMP_FOREIGN_MANUFACTURER_ADDRESS",
    "SECPA NUMBER":                                 "GMP_SECPA_NUMBER",
    "CERTIFICATE NUMBER":                           "GMP_CERTIFICATE_NUMBER",
    "TYPE OF ISSUANCE":                             "GMP_TYPE_OF_ISSUANCE",
    "CERTIFICATE VALIDITY":                         "GMP_CERTIFICATE_VALIDITY",
    "DECISION":                                     "GMP_DECISION",
    "STATUS":                                       "GMP_APP_STATUS",
    "RELEASED DATE":                                "GMP_RELEASED_DATE",
    "PROCESSED TIME":                               "GMP_PROCESSED_TIME",
    "END DATE":                                     "GMP_END_DATE",
    "TIMELINE":                                     "GMP_TIMELINE",
    "REMARKS":                                      "GMP_REMARKS",
    "1st DATE OF NOD":                              "GMP_NOD_DATE_1",
    "2nd DATE OF NOD":                              "GMP_NOD_DATE_2",
    "3rd DATE OF NOD":                              "GMP_NOD_DATE_3",
    "4th DATE OF NOD":                              "GMP_NOD_DATE_4",
    "5th DATE OF NOD":                              "GMP_NOD_DATE_5",
    "DATE PRINTED":                                 "GMP_DATE_PRINTED",
    "COMPLIANCE / ADDITIONAL DOCS DATE RECEIVED":   "GMP_COMPLIANCE_DOCS_DATE_RECEIVED",
    "PRODUCT LINE":                                 "GMP_PRODUCT_LINE",
}

# ─────────────────────────────────────────────────────────────────────────────
# Delegation log steps — same tuple format as main_db.py LOG_STEPS
# (step_label, user_col, id_col, decision_col, remarks_col, date_col, thread_col, del_idx)
#
# Aligned to the live 7-step workflow (see app/crud/gmp_record.py's
# GMP_LOG_STEPS / GMP_ACTION_ROUTES):
#   - "Quality Evaluator" renamed to "Evaluator"
#   - "QA Supervisor" removed entirely (step/group deleted)
#   - the 2nd-pass "Decker" slot repurposed and renamed to "QA Admin"
# del_idx values now match GMP_LOG_STEPS exactly (Decking=1, Evaluator=2,
# Checker=3, QA Admin=4, [LRD Chief Admin=5 — no Excel columns exist for this
# step, pre-existing gap, not introduced by this change], OD Receiving=6,
# OD Releasing=7).
# ─────────────────────────────────────────────────────────────────────────────
GMP_LOG_STEPS_EXCEL = [
    (
        "Decking",
        "DECKER", "DECKER ID", "DECKER DECISION", "DECKER REMARKS",
        "DATE DECKED END", "DECKER DEL THREAD", 1,
    ),
    (
        "Evaluator",
        "EVALUATOR", "EVALUATOR ID", "EVALUATOR DECISION", "EVALUATOR REMARKS",
        "DATE EVAL END", "EVALUATOR DEL THREAD", 2,
    ),
    (
        "Checker",
        "CHECKER", "CHECKER ID", "CHECKER DECISION", "CHECKER REMARKS",
        "DATE CHECKER END", "CHECKER DEL THREAD", 3,
    ),
    (
        "QA Admin",
        "QA ADMIN", "QA ADMIN ID", "QA ADMIN DECISION", "QA ADMIN REMARKS",
        "DATE QA ADMIN END", "QA ADMIN DEL THREAD", 4,
    ),
    (
        "OD Receiving",
        "OD-RECEIVING", "OD-RECEIVING ID", "OD-RECEIVING DECISION", "OD-RECEIVING REMARKS",
        "DATE OD-RECEIVING END", "OD-RECEIVING DEL THREAD", 6,
    ),
    (
        "OD Releasing",
        "OD-RELEASING", "OD-RELEASING ID", "OD-RELEASING DECISION", "OD-RELEASING REMARKS",
        "DATE OD-RELEASING END", "OD-RELEASING DEL THREAD", 7,
    ),
]

# One color per step group (matches main_db.py STEP_COLORS pattern)
STEP_COLORS = [
    "FFF2CC",   # Decking      — yellow
    "D9EAD3",   # Evaluator    — green
    "CFE2F3",   # Checker      — blue
    "D9D2E9",   # QA Admin     — purple
    "FCE5CD",   # OD Receiving — orange
    "F4CCCC",   # OD Releasing — red/rose
]

# ─────────────────────────────────────────────────────────────────────────────
# Date parser (same as main_db.py, plus Excel serial-number date support)
# ─────────────────────────────────────────────────────────────────────────────
def _parse_date(value):
    """
    Accepts virtually anything that could represent a date and normalizes it
    to a datetime (callers then take .date() where a plain date is needed):
      - a datetime / pandas Timestamp: returned as-is
      - a numeric Excel date serial (e.g. a cell read as a raw number because
        it wasn't formatted as a date in the sheet): converted from Excel's
        1899-12-30 epoch instead of being silently dropped
      - a string in virtually any common format (MM/DD/YYYY, "Jan 5 2025",
        ISO, etc.): parsed with dateutil's fuzzy parser
    Returns None only when the value is genuinely empty/unparseable.
    """
    if pd.isna(value) or value is None or value == "":
        return None
    if isinstance(value, (datetime, pd.Timestamp)):
        return value
    if isinstance(value, (int, float, np.integer, np.floating)):
        try:
            return pd.to_datetime(float(value), unit="D", origin="1899-12-30").to_pydatetime()
        except Exception:
            return None
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None
        try:
            return dateutil_parser.parse(value, fuzzy=True)
        except Exception:
            return None
    return None


# ─────────────────────────────────────────────────────────────────────────────
# Build the downloadable template workbook
# ─────────────────────────────────────────────────────────────────────────────
def _build_template_workbook() -> io.BytesIO:
    # All column headers in order
    main_columns   = list(GMP_COLUMN_MAPPING.keys())
    log_columns    = []
    for (_, user_col, id_col, dec_col, rem_col, date_col, thread_col, _idx) in GMP_LOG_STEPS_EXCEL:
        log_columns += [user_col, id_col, dec_col, rem_col, date_col, thread_col]

    all_columns = main_columns + log_columns

    # One blank sample row
    template_data = {col: [""] for col in all_columns}
    df = pd.DataFrame(template_data)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Template")
        workbook  = writer.book
        worksheet = writer.sheets["Template"]

        main_col_count = len(main_columns)

        # ── Style main DB columns — subtle gray (matches main_db.py) ──────────
        main_fill = PatternFill(start_color="EFEFEF", end_color="EFEFEF", fill_type="solid")
        main_font = Font(bold=True)
        for col_idx in range(1, main_col_count + 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.fill      = main_fill
            cell.font      = main_font
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            worksheet.column_dimensions[get_column_letter(col_idx)].width = 22

        # Widen a few long main columns
        wide = {
            "NAME OF ESTABLISHMENT": 32,
            "ADDRESS": 36,
            "FOREIGN MANUFACTURER": 26,
            "FOREIGN MANUFACTURER ADDRESS": 32,
            "COMPLIANCE / ADDITIONAL DOCS DATE RECEIVED": 32,
        }
        for col_idx, header in enumerate(main_columns, start=1):
            if header in wide:
                worksheet.column_dimensions[get_column_letter(col_idx)].width = wide[header]

        # ── Style log step columns — color per step, 6 cols each ──────────────
        log_col_start = main_col_count + 1
        for step_idx, _ in enumerate(GMP_LOG_STEPS_EXCEL):
            color = STEP_COLORS[step_idx % len(STEP_COLORS)]
            fill  = PatternFill(start_color=color, end_color=color, fill_type="solid")
            bold  = Font(bold=True)
            for offset in range(6):
                col_idx = log_col_start + (step_idx * 6) + offset
                cell = worksheet.cell(row=1, column=col_idx)
                cell.fill      = fill
                cell.font      = bold
                cell.alignment = Alignment(horizontal="center", wrap_text=True)
                worksheet.column_dimensions[get_column_letter(col_idx)].width = 22

        # Freeze header row
        worksheet.freeze_panes = "A2"

        # ── Instructions sheet ────────────────────────────────────────────────
        instructions = pd.DataFrame({
            "Column Group": [
                "Main GMP Columns (gray)",
                "DECKER (yellow)",
                "EVALUATOR (green)",
                "CHECKER (blue)",
                "QA ADMIN (purple)",
                "OD-RECEIVING (orange)",
                "OD-RELEASING (red/rose)",
                "ID Columns",
                "Del Thread Values",
                "Date Format",
                "DTN Field",
            ],
            "Description": [
                "Columns DTN → PRODUCT LINE. Core GMP record fields.",
                "First Decker step — 6 columns: Name, ID, Decision, Remarks, Date, Del Thread",
                "Evaluator step — 6 columns",
                "Checker step — 6 columns",
                "QA Admin step — 6 columns",
                "OD-Receiving step — 6 columns",
                "OD-Releasing step — 6 columns",
                "e.g. 'DECKER ID' — numeric employee ID (e.g. 1001). Leave blank if unknown.",
                "Del Thread per step: 'Open' (in progress) or 'Close' (completed).",
                "Preferred format is YYYY-MM-DD (e.g. 2025-01-15), but other common formats "
                "and Excel date-formatted cells are auto-detected and normalized automatically.",
                "DTN must be a number. Duplicate DTNs will be skipped on upload.",
            ],
            "Note": [
                "All fields optional except where required by your workflow.",
                "A log row is only created if the Name column (e.g. DECKER) has a value.",
                "Same rule applies for all step columns.",
                "Same rule applies for all step columns.",
                "Same rule applies for all step columns.",
                "Same rule applies for all step columns.",
                "Same rule applies for all step columns.",
                "Leave blank if no employee ID. Must be a whole number if filled.",
                "Defaults to 'Open' if left blank.",
                "Leave blank if not applicable.",
                "Leave blank if no DTN assigned yet.",
            ],
        })
        instructions.to_excel(writer, index=False, sheet_name="Instructions")

        # Style instructions header
        inst_ws   = writer.sheets["Instructions"]
        inst_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        inst_font = Font(bold=True, color="FFFFFF")
        for col_idx in range(1, 4):
            cell = inst_ws.cell(row=1, column=col_idx)
            cell.fill      = inst_fill
            cell.font      = inst_font
            cell.alignment = Alignment(horizontal="center")
        inst_ws.column_dimensions["A"].width = 32
        inst_ws.column_dimensions["B"].width = 55
        inst_ws.column_dimensions["C"].width = 45

    output.seek(0)
    return output
